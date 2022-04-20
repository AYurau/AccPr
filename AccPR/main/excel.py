import openpyxl
from openpyxl import Workbook

from .models import Product


def create_report(products, paymentType, date,DIR_NAME):
    day = str(date.day)
    try:
        book = openpyxl.open('./files/' + DIR_NAME + '/' + day + '.xlsx', read_only=False)
        sheet = book.active
        print(products, paymentType)
        for row in range(1, sheet.max_row + 2):
            if sheet[row][0].value is None:
                for items in products:
                    sheet[row][0].value = items['code']
                    sheet[row][1].value = items['name']
                    sheet[row][2].value = items['amount']
                    sheet[row][3].value = items['purchase']
                    sheet[row][4].value = items['price']
                    sheet[row][5].value = items['price'] - items['purchase']
                    sheet[row][6].value = paymentType
                    row += 1
        else:
            book.save(('./files/' + DIR_NAME + '/' + day + '.xlsx'))
            book.close()
    except:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Штрихкод'
        ws['B1'] = 'Наименование'
        ws['C1'] = 'Кол-во'
        ws['D1'] = 'Закупка'
        ws['E1'] = 'Розница'
        ws['F1'] = 'Заработок'
        ws['G1'] = 'Способ оплаты'
        for row in range(1, ws.max_row + 2):
            if ws[row][0].value is None:
                for items in products:
                    ws[row][0].value = items['code']
                    ws[row][1].value = items['name']
                    ws[row][2].value = items['amount']
                    ws[row][3].value = items['purchase']
                    ws[row][4].value = items['price']
                    ws[row][5].value = items['price'] - items['purchase']
                    ws[row][6].value = paymentType
                    row += 1
        else:
            wb.save('./files/' + DIR_NAME + '/' + day + '.xlsx')
            print('Файл создан. Значения записаны.')


def read_order():
    book = openpyxl.open('./files/order.xlsx', read_only=True)
    sheet = book.active
    for row in range(12, sheet.max_row + 1):
        code = sheet[row][5].value
        title = sheet[row][7].value
        count = sheet[row][20].value
        cost = sheet[row][22].value
        mrc = sheet[row][25].value
        if type(count) is int:
            if count > 0:
                products = Product.objects.all()
                for item in products:
                    if item.title == title:
                        print(f'Модель {item.title} существует. Добавлено {item.count} шт.')
                        item.count = item.count + count
                        item.save()
                        break
                else:
                    product = Product()
                    product.code = code
                    product.title = title
                    product.count = count
                    product.cost_fs = mrc
                    product.cost_m = cost
                    product.profit = cost - mrc
                    product.save()
                    print(f'Модель создана {product.title}')

    else:
        book.close()
        print('Работа завершена!')


def delete_products():
    i = 0
    products = Product.objects.all()
    for product in products:
        i += 1
        product.delete()
        print(f'{product.title} удалён!')
    else:
        print(f'Удалено {i} товаров')


def remains_model():
    book = openpyxl.open('./files/remains.xlsx', read_only=True)
    sheet = book.active
    for row in range(2, sheet.max_row + 1):
        product = Product()
        product.code = sheet[row][0].value
        product.title = sheet[row][1].value
        product.count = sheet[row][2].value
        product.cost_fs = sheet[row][3].value
        product.cost_m = sheet[row][4].value
        product.profit = int(product.cost_fs) - int(product.cost_m)
        product.save()
    else:
        book.close()
        print('Модели созданы!')

# def invoice_check():
#   inv = openpyxl.open('./files/remains.xlsx', read_only=False)
#  sheet = inv.active
# for titles in range(2, sheet.max_row+1):
#    title = sheet[titles][1].value
#   ttn = openpyxl.open('./files/ttn.xlsx', read_only=True)
#  sheets = ttn.active
# for ttn_rows in range(1, sheets.max_row+1):
#    if title == sheets[ttn_rows][0].value:
#       cost = sheets[ttn_rows][24].value
#      print(f'Добавлена оптовая цена {cost} для {title}')
#     sheet[titles][3].value = cost
#    sheet[titles][5].value = sheet[titles][4].value - cost
#   inv.save('./files/remains.xlsx')
#  inv.close()
# break
# else:
#   ttn.close()
# else:
#   print('Данные заявки и накладной скорректированы!')
